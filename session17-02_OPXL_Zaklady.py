'''
Openpyxl treba nainstalovat do obalky venv pomocou pipu

Umoznuje nam vytvarat z pythonu Excelovske tabulky, upravovat ich a nacitavat udaje do pythonu
'''


from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, Series
import random

if __name__ == '__main__':
    wb = Workbook()

    # WORKSHEET 1
    ws1 = wb.active
    ws1.title = 'NAS NAZOV'
    for row in range(1,12):
        ws1.append(range(12))

    # WORKSHEET 2
    ws2 = wb.create_sheet(title='MAIN')
    ws2['A1'] = 543435
    ws2['A2'] = 4334
    ws2['A3'] = -770
    ws2['A4'] = '=SUM(A1:A3)'  # formula

    # WORKSHEET 3
    ws3 = wb.create_sheet(title='DATA')
    for row in range(1,15):
        for col in range(1,15):
            _ = ws3.cell(column=col, row=row, value=random.randint(0,9999))
    print(ws3['B3'].value)
    print(ws3['B3'].number_format)  # Generak
    ws3.cell(row=3, column=2, value='SIGNAL LOST')

    alphabet = ['A','B','C','D']  # list
    for col in alphabet:
        cell = col+'15' # string reprezentujuci bunku
        formula = '=SUM(%s1:%s14)' % (col, col)
        ws3[cell] = formula

    ws3.merge_cells('A1:A4')
    ws3.unmerge_cells('A1:A4')

    ws3.insert_rows(5, 2)  # kde a kolko riadkov
    ws3.insert_cols(5, 2)  # kde a kolko riadkov




    # UKLADANIE DOKUMENTU
    wb.save(filename='cvicny.xlsx')
    print('SUBOR BOL ULOZENY')